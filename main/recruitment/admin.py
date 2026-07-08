import csv
import io
import json
import logging
from datetime import date, datetime

from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html

from .models import StaffApplication, StaffApplicationDocument

logger = logging.getLogger(__name__)

class StaffApplicationDocumentInline(admin.TabularInline):
    model = StaffApplicationDocument
    extra = 1
    readonly_fields = ("uploaded_at",)


EXPORT_FIELDS = [
    "id", "post_applied_for", "first_name", "middle_name", "surname",
    "father_name", "spouse_name", "category", "mobile_number", "email",
    "correspondence_address_line1", "correspondence_address_line2",
    "permanent_address_line1", "permanent_address_line2",
    "dob", "marital_status", "date_of_marriage",
    "education", "experience", "extra_curricular", "why_suitable",
    "epf_member", "epf_number", "references",
    "declaration_accepted", "declaration_date",
    "created_at", "updated_at",
]


def _serialize_value(value, *, for_excel: bool = False):
    if value is None:
        return "" if not for_excel else None
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if for_excel else value.isoformat()
    if isinstance(value, date):
        return value if for_excel else value.isoformat()
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def _get_export_rows(queryset):
    for app in queryset.order_by("-created_at"):
        row = []
        for field in EXPORT_FIELDS:
            value = getattr(app, field, None)
            row.append(value)
        yield row


def export_applications_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="staff_applications.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(EXPORT_FIELDS)

    for raw_row in _get_export_rows(queryset):
        writer.writerow([_serialize_value(v, for_excel=False) for v in raw_row])

    return response

export_applications_csv.short_description = "Download selected applications as CSV"


def export_applications_excel(modeladmin, request, queryset):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl is not installed — Excel export unavailable")
        return HttpResponse(
            "Excel export requires the openpyxl package. "
            "Install it with: pip install openpyxl",
            status=500,
            content_type="text/plain",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Applications"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    headers = [f.replace("_", " ").title() for f in EXPORT_FIELDS]
    ws.append(headers)

    for col_idx, _header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for raw_row in _get_export_rows(queryset):
        serialized = []
        for v in raw_row:
            sv = _serialize_value(v, for_excel=True)
            if isinstance(sv, str) and len(sv) > 32767:
                sv = sv[:32764] + "..."
            serialized.append(sv)
        ws.append(serialized)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = cell_alignment

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="staff_applications.xlsx"'
    return response

export_applications_excel.short_description = "Download selected applications as Excel"


@admin.register(StaffApplication)
class StaffApplicationAdmin(admin.ModelAdmin):
    list_display = (
        "full_name_display",
        "post_applied_for",
        "email",
        "mobile_number",
        "category",
        "marital_status",
        "created_at",
    )
    search_fields = ("first_name", "surname", "email", "mobile_number", "post_applied_for")
    list_filter = ("post_applied_for", "category", "marital_status", "epf_member", "created_at")
    readonly_fields = ("created_at", "updated_at", "full_name_display")
    actions = [export_applications_csv, export_applications_excel]
    inlines = [StaffApplicationDocumentInline]

    fieldsets = (
        (
            "Basic Details",
            {
                "fields": (
                    "post_applied_for",
                    "surname",
                    "first_name",
                    "middle_name",
                    "father_name",
                    "spouse_name",
                    "category",
                    "photograph",
                )
            },
        ),
        (
            "Contact Details",
            {
                "fields": (
                    "correspondence_address_line1",
                    "correspondence_address_line2",
                    "permanent_address_line1",
                    "permanent_address_line2",
                    "mobile_number",
                    "email",
                )
            },
        ),
        (
            "Personal Details",
            {
                "fields": (
                    "dob",
                    "marital_status",
                    "date_of_marriage",
                )
            },
        ),
        (
            "Education",
            {
                "fields": ("education",),
                "description": (
                    "JSON array. Each entry: examination_name, school_college, "
                    "board_university, year_of_passing, medium, division, percentage."
                ),
            },
        ),
        (
            "Experience",
            {
                "fields": ("experience",),
                "description": (
                    "JSON array. Each entry: organization, designation, "
                    "from_date (YYYY-MM-DD), to_date, job_profile, last_salary."
                ),
            },
        ),
        (
            "Additional Information",
            {
                "fields": (
                    "extra_curricular",
                    "why_suitable",
                    "epf_member",
                    "epf_number",
                )
            },
        ),
        (
            "References",
            {
                "fields": ("references",),
                "description": "JSON array. Each entry: name, address, contact_number.",
            },
        ),
        (
            "Declaration",
            {
                "fields": (
                    "declaration_accepted",
                    "declaration_date",
                )
            },
        ),
        (
            "Attachments",
            {
                "fields": (
                    "attach_10th",
                    "attach_12th",
                    "attach_diploma",
                    "attach_graduation",
                    "attach_post_graduation",
                    "attach_pan_card",
                    "attach_aadhaar",
                    "attach_form16",
                    "attach_last_salary",
                    "attach_experience_certs",
                    "attach_fitness_cert",
                    "attach_photos",
                ),
                "classes": ("collapse",),
            },
        ),
        (
            "Meta",
            {
                "fields": ("created_at", "updated_at"),
                "classes": ("collapse",),
            },
        ),
    )

    @admin.display(description="Full Name")
    def full_name_display(self, obj: StaffApplication) -> str:
        return obj.full_name

@admin.register(StaffApplicationDocument)
class StaffApplicationDocumentAdmin(admin.ModelAdmin):
    list_display = ("application_name", "document_type", "file_link", "uploaded_at")
    list_filter = ("document_type", "uploaded_at")
    search_fields = (
        "application__first_name",
        "application__surname",
        "application__email",
    )
    readonly_fields = ("uploaded_at",)

    @admin.display(description="Applicant")
    def application_name(self, obj: StaffApplicationDocument) -> str:
        return obj.application.full_name

    @admin.display(description="File")
    def file_link(self, obj: StaffApplicationDocument) -> str:
        if obj.file:
            return format_html('<a href="{}" target="_blank">View</a>', obj.file.url)
        return "—"